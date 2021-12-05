from django.shortcuts import redirect, render
from django.http import HttpResponse #################
from .models import Habit, end_Habit ################
import json
import datetime
import openpyxl
from . import dateRange
import pandas
import matplotlib.pyplot
from calendar import HTMLCalendar
from . import HTMLCalendar
import calendar
from . import statistic_html
import pytz
import sys
import os

def home(request): #함수로 구현
    return render(request, 'hb_main/home.html')

def result(request): 
    name = request.POST.get('name')  #######################################
    utc=pytz.UTC
    habits = Habit.objects.all()
    KST = pytz.timezone('Asia/Seoul')
    for i in habits:
        end=i.start_day+datetime.timedelta(days=65)
        #utc.localize(end)
        td=datetime.datetime.now()
        utc.localize(td)
        #datetime.datetime.today()=utc.localize(datetime.datetime.today())
        if end<=utc.localize(td): #습관의 마지막 날짜보다 오늘 날짜가 크면,
            i.end=True #객체의 end값을 True로 변경
            i.save()
        if i.end==True: #end가 True가 된 객체들을,
            end_habit=end_Habit.objects.create(name=i.name,start_day=i.start_day,do_it_list=i.do_it_list)
            end_habit.save() #end_Habit 모델로 추가(66일이 지나서 끝난 습관 모델)
            i.delete() #Habit 모델에서는 삭제(습관 모델)
    end_habits=end_Habit.objects.all()
    context={'habits':habits, 'end_habits':end_habits} 
    return render(request, 'hb_main/result.html', context)
    #habit = Habit.objects.all()
    #return HttpResponse("asdf")

def adding(request): 
    if request.method == 'POST':
        #return HttpResponse(request)
        habit=Habit.objects.create(name=request.POST['name'])
        dates=list()
        habit.do_it_list=json.dumps(dates) #객체의 do_it_list 필드에는 리스트 형태로 날짜 저장
        habit.save()
        return render(request, 'hb_main/home.html')
    return render(request, 'hb_main/adding.html')
    
def doit(request): #*
    if request.method == 'POST':
        select = request.POST['true']
        Doit=Habit.objects.get(id=select)
        Doit.do_it=True
        jsonDec=json.decoder.JSONDecoder()
        dates=jsonDec.decode(Doit.do_it_list)
        dates.append((datetime.datetime.today()).strftime("%Y-%m-%d"))
        Doit.do_it_list=json.dumps(dates)
        Doit.save()
        return render(request,'hb_main/doit.html')

def statistics(request, habit_id): #*
    if request.method == 'GET':
        select = request.GET.get('habit_id')
        statistic=Habit.objects.get(id=habit_id)
        context={'statistic':statistic}
        jsonDec=json.decoder.JSONDecoder() #데이터베이스에 있는 실천 여부 정보 리스트로 받아오기
        dates=jsonDec.decode(statistic.do_it_list)
        wb=openpyxl.Workbook() #엑셀에 정보 추가
        file='./day.xlsx'
        wb.active.title="habit_day_66"
        w1=wb["habit_day_66"]
        end_day=statistic.start_day + datetime.timedelta(days=65)
        days=dateRange.date_list(statistic.start_day,end_day)
        w1.cell(1,1).value="date"
        w1.cell(1,2).value="action"
        for i in range(2,len(days)+2):
            w1.cell(i,1).value=days[i-2] #날짜 추가
        get_cells = w1['A2':'A67']
        num=1
        for i in get_cells:
            for j in i:
                num+=1
                for n in dates:
                    if j.value==n:
                        w1.cell(num,2).value="True" #실천 여부 추가
        wb.save(file)
        html_file=statistic_html.make_html()
        dir = os.path.dirname(os.path.realpath('day.xlsx'))
        temp = dir + '\\day.xlsx'
        data=pandas.read_excel(temp) #엑셀 데이터 읽기
        # data=pandas.read_excel("C:\\Users\\subin\\Desktop\\habitt\\6ha6it\\wp_pj8\\habit\\day.xlsx") #엑셀 데이터 읽기
        data=pandas.DataFrame(data)
        htmlCalendar = HTMLCalendar.HTMLCalendar(calendar.SUNDAY)
        month_action=list() #엑셀에 있는 날짜 문자열에서 월 정수형 정보 중복없이 가져오기
        for m in data.date:
            month_action.append(int(m[5:7]))
        month_set=set(month_action)
        month_action=list(month_set)
        html_file=statistic_html.month_html(htmlCalendar, sorted(month_action), data)
        html_file=statistic_html.day_html(data)
        html_file=statistic_html.week_html(data)
        return render(request, 'hb_main/statistics.html', context)

def deleting(request):
    if request.method == "POST":
        if request.POST.get('del_habit')!=None:
            habit=request.POST.get('del_habit')
            del_habit=end_Habit.objects.get(name=habit)
            del_habit.delete()
            return render(request, 'hb_main/deleting.html')
        else:
            habit=request.POST.get('del_habit2')
            del_habit=Habit.objects.get(name=habit)
            del_habit.delete()
            return render(request, 'hb_main/deleting.html')

def statistics2(request, habit_id):
    if request.method == 'GET':
        select = request.GET.get('habit_id')
        statistic=end_Habit.objects.get(id=habit_id)
        context={'statistic':statistic}
        jsonDec=json.decoder.JSONDecoder() #데이터베이스에 있는 실천 여부 정보 리스트로 받아오기
        dates=jsonDec.decode(statistic.do_it_list)
        wb=openpyxl.Workbook() #엑셀에 정보 추가
        file='./day.xlsx'
        wb.active.title="habit_day_66"
        w1=wb["habit_day_66"]
        end_day=statistic.start_day + datetime.timedelta(days=65)
        days=dateRange.date_list(statistic.start_day,end_day)
        w1.cell(1,1).value="date"
        w1.cell(1,2).value="action"
        for i in range(2,len(days)+2):
            w1.cell(i,1).value=days[i-2] #날짜 추가
        get_cells = w1['A2':'A67']
        num=1
        for i in get_cells:
            for j in i:
                num+=1
                for n in dates:
                    if j.value==n:
                        w1.cell(num,2).value="True" #실천 여부 추가
        wb.save(file)
        html_file=statistic_html.make_html()
        dir = os.path.dirname(os.path.realpath('day.xlsx'))
        temp = dir + '\\day.xlsx'
        data=pandas.read_excel(temp) #엑셀 데이터 읽기
        # data=pandas.read_excel("\\6ha6it\\wp_pj8\\habit\\day.xlsx") #엑셀 데이터 읽기
        data=pandas.DataFrame(data)
        htmlCalendar = HTMLCalendar.HTMLCalendar(calendar.SUNDAY)
        month_action=list() #엑셀에 있는 날짜 문자열에서 월 정수형 정보 중복없이 가져오기
        for m in data.date:
            month_action.append(int(m[5:7]))
        month_set=set(month_action)
        month_action=list(month_set)
        html_file=statistic_html.month_html(htmlCalendar, sorted(month_action), data)
        html_file=statistic_html.day_html(data)
        html_file=statistic_html.week_html(data)
        return render(request, 'hb_main/statistics.html', context)


def updating(request, habit_id=None):
    global up_habit
    if request.method == "GET":
        #temp=request.POST.get(habit_id)
        up_habit=Habit.objects.get(id=habit_id)
        up_habit=habit_id
        #up=Habit.objects.get(name=temp)
        #return HttpResponse(up_habit)
    if request.method == "POST":
        habit=request.POST.get('up_habit')
        up=Habit.objects.get(id=up_habit)
        habit=Habit.objects.filter(id=up_habit)
        #return HttpResponse(up.name)
        if request.POST.get('name')!="":
            habit.update(name=request.POST.get('name'))
        if request.POST.get('start_day')!="":
            habit.update(start_day=request.POST.get('start_day'))
            
        if request.POST.get('action_day')!="":
            do=request.POST.get('action_day')
            #return HttpResponse(up.do_it_list)
            jsonDec=json.decoder.JSONDecoder()
            dates=jsonDec.decode(up.do_it_list)
            dates.append(do)
            habit.update(do_it_list=json.dumps(dates))
            #return HttpResponse(type(dates))
        if request.POST.get('no_action_day')!="":
            no_do=request.POST.get('no_action_day')
            jsonDec=json.decoder.JSONDecoder()
            dates=jsonDec.decode(up.do_it_list)
            dates.remove(no_do)
            habit.update(do_it_list=json.dumps(dates))
        return render(request, 'hb_main/home.html')
    return render(request, 'hb_main/updating.html')